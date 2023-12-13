import matplotlib.pyplot as plt
import pandas as pd
import os
import os.path
import openpyxl

def read_xlsx_as_rows(file_path):
    """Создаем датафрэйм их экселя"""
    try:
        df = pd.read_excel(file_path)
        return df
    except Exception as e:
        print(f"An error occurred: {e}")
        return None

# Пример использования:
# Конвертируем эксель в датафрэйм
file_path = 'output_myvar.xlsx'
result_df = read_xlsx_as_rows(file_path)

# удаляем результирующий файл, если он есть
if os.path.isfile('output_df_xlsx.xlsx'):
    os.remove('output_df_xlsx.xlsx')

# ВЫбираем уникальные значения из столбцов
strana = result_df['strana'].unique()
region = result_df['region'].unique()
# выводим заголовок, 0 строку, на печать в виде списка
unique_tolist = result_df.columns.tolist()
print(strana[0])
print(region[0])
print(unique_tolist)
# вывод строки по индексу с 0, где заголовок метка
# print(result_df.loc[[2]])

# получаем список стран использующий ветрогенерацию
wind_strana = [i[2] for i in result_df.loc[result_df['energe'] == 'WindGeneration-TWh'].values.tolist()]

# перебираем список регионов, 2 столбец
for i in region:
    # перебираем список стран, 3 столбец
    for k in strana:
        # проверяем данная страна производит ветрогенерацию или нет
        if k in wind_strana:
            # сохраняем список индексов для страны по всему датафрэйму
            indices = result_df.loc[(result_df['region'] == i) & (result_df['strana'] == k)].index
            # вводим переменную, с которой начнём вводить данные в новый эксель
            indstr = 2
            # print(indices)
            # перебираем все индексы
            for n in indices:
                print(result_df.loc[[n]].values.tolist()[0][2], result_df.loc[[n]].values.tolist()[0][1],
                result_df.loc[[n]].values.tolist()[0][0])
                # проверяем наличие файла, если его нету, создаем иначе открываем
                if not os.path.isfile('output_df_xlsx.xlsx'):
                    wb = openpyxl.Workbook()
                else:
                    wb = openpyxl.load_workbook('output_df_xlsx.xlsx')
                # print(wb.sheetnames)
                # проверяем, лист с именем страны
                if result_df.loc[[n]].values.tolist()[0][2] not in wb.sheetnames:
                    # создаем лист
                    wb.create_sheet(title=f'{result_df.loc[[n]].values.tolist()[0][2]}')
                    # переключаемся на этот лист
                    sheet = wb[f'{result_df.loc[[n]].values.tolist()[0][2]}']
                    # в первую ячейку вводим страну
                    sheet.cell(row=1, column=1, value=f'{result_df.loc[[n]].values.tolist()[0][2]}')
                    # во вторую ячейку - регион
                    sheet.cell(row=1, column=2, value=f'{result_df.loc[[n]].values.tolist()[0][1]}')
                    # в следующую строку год и в остальные ячейки этой строки значения каждого года
                    sheet.cell(row=indstr, column=1, value='Год')
                    b = 2
                    for m in unique_tolist[3:]:
                        sheet.cell(row=indstr, column=b, value=m)
                        b += 1
                # активируем лист с страной
                sheet = wb[f'{result_df.loc[[n]].values.tolist()[0][2]}']
                # переходим к следующей строке
                indstr += 1
                # циклически в строку вводим вид генерации и их значения
                sheet.cell(row=indstr, column=1, value=f'{result_df.loc[[n]].values.tolist()[0][0]}')
                b = 2
                for m in result_df.loc[[n]].values.tolist()[0][3:]:
                    sheet.cell(row=indstr, column=b, value=m)
                    b += 1
                # переходим к следующей строке
                indstr += 1


                # Тестируем вставку изображения
                list1 = unique_tolist[3:]
                list2 = result_df.loc[[n]].values.tolist()[0][3:]
                # Находим индекс последнего 0 значения во втором списке
                try:
                    index = 0
                    for idx, value in enumerate(list2, start=1):
                        if value == 0:
                            index = idx
                except BaseException:
                    print("Не получилось узнать индекс")

                # print(index)
                # Генерируем график 1
                plt.figure()
                plt.plot(list1, list2)
                plt.xlabel('Год')
                plt.ylabel('Генерация')
                plt.title(f'{result_df.loc[[n]].values.tolist()[0][0]}')
                # Создаем путь к временному файлу с изображением графика
                graph_filename = 'graph_temp1.png'
                # Сохраняем график как изображение
                plt.savefig(graph_filename, dpi=70)  # dpi - точек на дюйм
                plt.close()
                # Генерируем график 2
                plt.figure()
                plt.plot(list1[index:], list2[index:])
                plt.xlabel('Год')
                plt.ylabel('Генерация')
                plt.title(f'{result_df.loc[[n]].values.tolist()[0][0]} без учета 0 показателей')
                # Создаем путь к временному файлу с изображением графика
                graph_filename2 = 'graph_temp2.png'
                # Сохраняем график как изображение
                plt.savefig(graph_filename2, dpi=70)  # dpi - точек на дюйм
                plt.close()

                # Создаем объект изображения
                img = openpyxl.drawing.image.Image(graph_filename)
                img2 = openpyxl.drawing.image.Image(graph_filename2)
                # Вставляем изображение 1 в указанные координаты
                cell = sheet.cell(row=indstr, column=1)
                cell.coordinate
                sheet.add_image(img, cell.coordinate)
                # Вставляем изображение 2 в указанные координаты
                cell = sheet.cell(row=indstr, column=8)
                cell.coordinate
                sheet.add_image(img2, cell.coordinate)

                indstr += 16

                # Сохраняем и закрываем книгу
                wb.save('output_df_xlsx.xlsx')
                wb.close()
                # Удаляем временное изображение
                os.remove(graph_filename)
                os.remove(graph_filename2)

                # break
            # break


